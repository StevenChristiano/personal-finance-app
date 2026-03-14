import calendar
import io
from datetime import datetime, timezone
from typing import Optional, List

import numpy as np
import openpyxl
import pandas as pd
from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from auth import get_current_user
from database import Category, Transaction, User, UserModel, get_db
from services.ml import (
    THRESHOLD_ANOMALY, THRESHOLD_WARNING,
    calculate_score, global_encoder, global_model,
    global_norm_params, global_scaler, load_user_model,
    retrain_user_model, should_retrain,
)

router = APIRouter(tags=["transactions"])


# ============================================================
# SCHEMAS
# ============================================================
class TransactionCreate(BaseModel):
    amount: float = Field(..., gt=0)
    category_id: int
    note: Optional[str] = None
    timestamp: Optional[str] = None


class BulkTransactionItem(BaseModel):
    amount: float = Field(..., gt=0)
    category_id: int
    note: Optional[str] = None
    timestamp: Optional[str] = None


class BulkTransactionSave(BaseModel):
    transactions: List[BulkTransactionItem]


# ============================================================
# HELPERS
# ============================================================
def _get_model(user_id: int, db: Session):
    """Return model + artifacts untuk user, fallback ke global model."""
    user_model_row = db.query(UserModel).filter(
        UserModel.user_id == user_id,
        UserModel.is_trained == True,
    ).first()
    if user_model_row:
        return load_user_model(user_model_row)
    if global_model:
        return global_model, global_scaler, global_encoder, global_norm_params
    return None, None, None, None


def _score_transaction(amount, category_name, ts, model, scaler, encoder, norm_params,
                        warning_threshold, anomaly_threshold):
    """Return (score, status) atau (None, None) kalau model tidak tersedia."""
    if model is None or category_name not in encoder.classes_:
        return None, None
    try:
        return calculate_score(
            amount, category_name, ts, model, scaler, encoder, norm_params,
            warning_threshold=warning_threshold,
            anomaly_threshold=anomaly_threshold,
        )
    except Exception:
        return None, None


def _build_excel_template(categories: list) -> bytes:
    """Buat Excel template dan return sebagai bytes."""
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Transactions"

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    # Header row
    header_fill  = PatternFill("solid", fgColor="1A1A1A")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    for col_idx, (header, width) in enumerate(
        zip(["date", "amount", "category", "note"], [22, 18, 22, 30]), start=1
    ):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Example rows
    now_str      = datetime.now().strftime("%Y-%m-%d %H:%M")
    example_fill = PatternFill("solid", fgColor="F9FAFB")
    for row_idx, (amount, cat, note) in enumerate([
        (50000, "Food", "Lunch with team"),
        (25000, "Transport", "Gojek to office"),
        (150000, "Entertainment", "Cinema ticket"),
    ], start=2):
        for col_idx, value in enumerate([now_str, amount, cat, note], start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = example_fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")

    # Category dropdown validation
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(c.name for c in categories)}"',
        allow_blank=True,
    )
    dv.sqref = "C2:C10000"
    ws.add_data_validation(dv)
    ws.freeze_panes = "A2"

    # Reference sheet
    ws2 = wb.create_sheet("Category Reference")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 50
    ref_fill = PatternFill("solid", fgColor="1A1A1A")
    for ci, hdr in enumerate(["Category Name", "Anomaly Scan", "Description"], start=1):
        cell = ws2.cell(row=1, column=ci, value=hdr)
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.fill      = ref_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws2.row_dimensions[1].height = 22

    descriptions = {
        "Food": "Daily food & beverage expenses",
        "Transport": "Commuting and travel costs",
        "Lifestyle": "Personal care and clothing",
        "Entertainment": "Leisure activities",
        "Utilities": "Household bills",
        "Telecommunication": "Phone-related expenses",
        "Subscription": "Recurring digital services",
        "Health": "Medical expenses (Not monitored)",
        "Education": "Learning costs (Not monitored)",
        "Big Expense": "Large one-time purchases (Not monitored)",
    }
    for ri, cat in enumerate(categories, start=2):
        ws2.cell(row=ri, column=1, value=cat.name).border = thin_border
        scan = ws2.cell(row=ri, column=2, value="No" if cat.is_excluded else "Yes")
        scan.border    = thin_border
        scan.font      = Font(color="DC2626" if cat.is_excluded else "16A34A", bold=True)
        scan.alignment = Alignment(horizontal="center")
        ws2.cell(row=ri, column=3, value=descriptions.get(cat.name, "")).border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_upload_row(row, df_columns, categories: dict):
    """Parse satu row dari uploaded file. Return (amount, category_obj, ts, note, errors)."""
    errors = []

    # Amount
    try:
        amount = float(row["amount"])
        if amount <= 0:
            errors.append("Amount must be > 0")
    except (ValueError, TypeError):
        amount = 0.0
        errors.append("Invalid amount")

    # Category
    raw_cat      = str(row["category"]).strip() if pd.notna(row.get("category")) else ""
    category_obj = categories.get(raw_cat.lower())
    if not category_obj:
        errors.append(f"Unknown category: '{raw_cat}'")

    # Timestamp
    try:
        raw_date = row.get("date", "")
        if pd.isna(raw_date):
            ts = datetime.now()
        else:
            ts = pd.to_datetime(raw_date).replace(tzinfo=None).to_pydatetime()
    except (ValueError, TypeError):
        ts = datetime.now()
        errors.append("Invalid date format — defaulting to now")

    # Note
    note = str(row["note"]).strip() if "note" in df_columns and pd.notna(row.get("note")) else None

    return amount, category_obj, ts, note, errors


# ============================================================
# ENDPOINTS
# ============================================================

# Categories Dropdown
@router.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    return [{"id": c.id, "name": c.name, "is_excluded": c.is_excluded}
            for c in db.query(Category).all()]

# Create Transaction
@router.post("/transactions", status_code=201)
def create_transaction(
    req: TransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    category = db.query(Category).filter(Category.id == req.category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category Not Found.")

    ts                = datetime.fromisoformat(req.timestamp) if req.timestamp else datetime.now(timezone.utc)
    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY

    anomaly_score, anomaly_status = None, None
    if not category.is_excluded:
        model, scaler, encoder, norm_params = _get_model(current_user.id, db)
        anomaly_score, anomaly_status = _score_transaction(
            req.amount, category.name, ts,
            model, scaler, encoder, norm_params,
            warning_threshold, anomaly_threshold,
        )

    transaction = Transaction(
        user_id        = current_user.id,
        category_id    = req.category_id,
        amount         = req.amount,
        note           = req.note,
        timestamp      = ts,
        anomaly_score  = anomaly_score,
        anomaly_status = anomaly_status,
        is_excluded    = category.is_excluded,
    )
    db.add(transaction)
    db.commit()
    db.refresh(transaction)

    if should_retrain(current_user.id, db):
        retrain_user_model(current_user.id, db, manual=False)

    return {
        "id"            : transaction.id,
        "amount"        : transaction.amount,
        "category"      : category.name,
        "note"          : transaction.note,
        "timestamp"     : transaction.timestamp,
        "anomaly_score" : transaction.anomaly_score,
        "anomaly_status": transaction.anomaly_status,
        "is_excluded"   : transaction.is_excluded,
        "message"       : (
            "⚠️ Anomaly Detected!" if anomaly_status == "anomaly"
            else "🔔 Unusual Transaction" if anomaly_status == "warning"
            else "✅ Normal Transaction."
        ),
    }

# List Transactions with optional month/year filter
@router.get("/transactions")
def get_transactions(
    month: Optional[int] = None,
    year: Optional[int] = None,
    warning_threshold: float = 0.50,
    anomaly_threshold: float = 0.60,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if month and year:
        last_day = calendar.monthrange(year, month)[1]
        query = query.filter(
            Transaction.timestamp >= datetime(year, month, 1),
            Transaction.timestamp <= datetime(year, month, last_day, 23, 59, 59),
        )
    transactions = query.order_by(Transaction.timestamp.desc()).all()
    categories   = {cat.id: cat.name for cat in db.query(Category).all()}

    def dynamic_status(score):
        if score is None:            return None
        if score >= anomaly_threshold: return "anomaly"
        if score >= warning_threshold: return "warning"
        return "normal"

    return [{
        "id"            : t.id,
        "amount"        : t.amount,
        "category_id"   : t.category_id,
        "category_name" : categories.get(t.category_id, "Unknown"),
        "note"          : t.note,
        "timestamp"     : t.timestamp,
        "anomaly_score" : t.anomaly_score,
        "anomaly_status": dynamic_status(t.anomaly_score) if not t.is_excluded else None,
        "is_excluded"   : t.is_excluded,
    } for t in transactions]

# Delete Transaction
@router.delete("/transactions/{transaction_id}")
def delete_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == current_user.id,
    ).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found.")
    db.delete(transaction)
    db.commit()
    return {"message": "Transaction Successfully Deleted."}

# Bulk Upload Template Download
@router.get("/transactions/template")
def download_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    categories = db.query(Category).all()
    content    = _build_excel_template(categories)
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="transaction_template.xlsx"'},
    )

# Bulk Upload Preview
@router.post("/transactions/upload-preview")
async def upload_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded.")

    filename_lower = file.filename.lower()
    if not any(filename_lower.endswith(ext) for ext in (".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted.")

    content = await file.read()
    buf     = io.BytesIO(content)
    try:
        df = (pd.read_csv(buf) if filename_lower.endswith(".csv")
              else pd.read_excel(buf, engine="openpyxl", sheet_name=0))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {e}")

    df.columns = [str(c).strip().lower() for c in df.columns]
    missing    = {"date", "amount", "category"} - set(df.columns)
    if missing:
        raise HTTPException(status_code=400, detail=f"Missing columns: {', '.join(missing)}")

    df = df.dropna(subset=["date", "amount", "category"], how="all").reset_index(drop=True)
    if df.empty:
        raise HTTPException(status_code=400, detail="No valid data rows.")

    categories        = {c.name.lower(): c for c in db.query(Category).all()}
    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY
    model, scaler, encoder, norm_params = _get_model(current_user.id, db)

    preview_rows = []
    for idx, row in df.iterrows():
        amount, category_obj, ts, note, errors = _parse_upload_row(row, df.columns, categories)
        is_excluded = category_obj.is_excluded if category_obj else False

        anomaly_score, anomaly_status = None, None
        if not errors and not is_excluded and category_obj:
            anomaly_score, anomaly_status = _score_transaction(
                amount, category_obj.name, ts,
                model, scaler, encoder, norm_params,
                warning_threshold, anomaly_threshold,
            )

        preview_rows.append({
            "_row"          : int(idx) + 2,
            "timestamp"     : ts.isoformat(),
            "amount"        : amount,
            "category_name" : category_obj.name if category_obj else str(row.get("category", "")),
            "category_id"   : category_obj.id if category_obj else None,
            "note"          : note,
            "anomaly_score" : anomaly_score,
            "anomaly_status": anomaly_status,
            "is_excluded"   : is_excluded,
            "errors"        : errors,
        })

    return {"rows": preview_rows, "total": len(preview_rows)}

# Bulk Upload Save
@router.post("/transactions/bulk-save", status_code=201)
def bulk_save(
    req: BulkTransactionSave,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    categories        = {c.id: c for c in db.query(Category).all()}
    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY
    model, scaler, encoder, norm_params = _get_model(current_user.id, db)

    saved = []
    for item in req.transactions:
        category = categories.get(item.category_id)
        if not category:
            continue

        ts            = datetime.fromisoformat(item.timestamp) if item.timestamp else datetime.now(timezone.utc)
        is_excluded   = category.is_excluded
        anomaly_score, anomaly_status = None, None

        if not is_excluded:
            anomaly_score, anomaly_status = _score_transaction(
                item.amount, category.name, ts,
                model, scaler, encoder, norm_params,
                warning_threshold, anomaly_threshold,
            )

        tx = Transaction(
            user_id        = current_user.id,
            category_id    = item.category_id,
            amount         = item.amount,
            note           = item.note,
            timestamp      = ts,
            anomaly_score  = anomaly_score,
            anomaly_status = anomaly_status,
            is_excluded    = is_excluded,
        )
        db.add(tx)
        saved.append(tx)

    db.commit()
    for tx in saved:
        db.refresh(tx)

    if should_retrain(current_user.id, db):
        retrain_user_model(current_user.id, db, manual=False)

    return {"saved": len(saved), "message": f"{len(saved)} transactions saved successfully."}