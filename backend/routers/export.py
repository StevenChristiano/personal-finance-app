import calendar
import io
from datetime import datetime
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Response
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
)
from sqlalchemy.orm import Session

from auth import get_current_user
from database import Category, Transaction, User, get_db
from services.ml import THRESHOLD_ANOMALY, THRESHOLD_WARNING

router = APIRouter(prefix="/transactions/export", tags=["export"])

MONTHS_ID = [
    "", "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember"
]

STATUS_LABEL = {
    "normal" : "Normal",
    "warning": "Warning",
    "anomaly": "Anomaly",
    None     : "-",
}


def _get_transactions(user_id, month, year, db, warning_threshold, anomaly_threshold):
    last_day = calendar.monthrange(year, month)[1]
    txs = db.query(Transaction).filter(
        Transaction.user_id    == user_id,
        Transaction.timestamp  >= datetime(year, month, 1),
        Transaction.timestamp  <= datetime(year, month, last_day, 23, 59, 59),
    ).order_by(Transaction.timestamp.asc()).all()
    categories = {c.id: c.name for c in db.query(Category).all()}

    def status(t):
        if t.is_excluded or t.anomaly_score is None: return None
        if t.anomaly_score >= anomaly_threshold: return "anomaly"
        if t.anomaly_score >= warning_threshold: return "warning"
        return "normal"

    return [
        {
            "date"    : t.timestamp.strftime("%d %b %Y %H:%M"),
            "category": categories.get(t.category_id, "-"),
            "amount"  : t.amount,
            "note"    : t.note or "-",
            "status"  : status(t),
            "score"   : round(t.anomaly_score * 100, 1) if t.anomaly_score is not None else None,
        }
        for t in txs
    ]


# ============================================================
# EXCEL EXPORT
# ============================================================
@router.get("/excel")
def export_excel(
    month: int,
    year: int,
    include_score: bool = False,
    warning_threshold: float = THRESHOLD_WARNING,
    anomaly_threshold: float = THRESHOLD_ANOMALY,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = _get_transactions(current_user.id, month, year, db, warning_threshold, anomaly_threshold)
    if not rows:
        raise HTTPException(status_code=404, detail="No transactions found for this period.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{MONTHS_ID[month]} {year}"

    thin = Border(
        left=Side(style="thin", color="E5E7EB"), right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),  bottom=Side(style="thin", color="E5E7EB"),
    )

    # Title
    ws.merge_cells("A1:F1" if include_score else "A1:E1")
    title_cell = ws["A1"]
    title_cell.value     = f"Laporan Transaksi — {MONTHS_ID[month]} {year}"
    title_cell.font      = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["Tanggal", "Kategori", "Amount (Rp)", "Note", "Status"]
    widths  = [22, 18, 18, 30, 12]
    if include_score:
        headers.append("Score (%)")
        widths.append(12)

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font      = Font(color="FFFFFF", bold=True, size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    # Status colors
    status_fills = {
        "normal" : PatternFill("solid", fgColor="DCFCE7"),
        "warning": PatternFill("solid", fgColor="FEF9C3"),
        "anomaly": PatternFill("solid", fgColor="FEE2E2"),
    }
    status_fonts = {
        "normal" : Font(color="166534", bold=True),
        "warning": Font(color="854D0E", bold=True),
        "anomaly": Font(color="991B1B", bold=True),
    }

    # Data rows
    total = 0
    for i, row in enumerate(rows, start=3):
        vals = [row["date"], row["category"], row["amount"], row["note"], STATUS_LABEL[row["status"]]]
        if include_score:
            vals.append(f"{row['score']}%" if row["score"] is not None else "-")

        for col, val in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border    = thin
            cell.alignment = Alignment(vertical="center")
            if col == 3:  # amount
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            if col == 5 and row["status"]:  # status
                cell.fill = status_fills.get(row["status"], PatternFill())
                cell.font = status_fonts.get(row["status"], Font())
                cell.alignment = Alignment(horizontal="center", vertical="center")
        total += row["amount"]
        ws.row_dimensions[i].height = 18

    # Total row
    total_row = len(rows) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=1).border = thin
    total_cell = ws.cell(row=total_row, column=3, value=total)
    total_cell.font         = Font(bold=True)
    total_cell.number_format = '#,##0'
    total_cell.alignment    = Alignment(horizontal="right", vertical="center")
    total_cell.border       = thin
    for col in [2, 4, 5] + ([6] if include_score else []):
        ws.cell(row=total_row, column=col).border = thin

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"transaksi_{MONTHS_ID[month].lower()}_{year}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# PDF EXPORT
# ============================================================
@router.get("/pdf")
def export_pdf(
    month: int,
    year: int,
    include_score: bool = False,
    warning_threshold: float = THRESHOLD_WARNING,
    anomaly_threshold: float = THRESHOLD_ANOMALY,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = _get_transactions(current_user.id, month, year, db, warning_threshold, anomaly_threshold)
    if not rows:
        raise HTTPException(status_code=404, detail="No transactions found for this period.")

    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4, leftMargin=15*mm, rightMargin=15*mm, topMargin=15*mm, bottomMargin=15*mm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=14, spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"],   fontSize=9,  textColor=colors.gray, spaceAfter=12)
    story.append(Paragraph(f"Laporan Transaksi", title_style))
    story.append(Paragraph(f"{MONTHS_ID[month]} {year}  •  {len(rows)} transaksi", sub_style))

    # Summary
    total   = sum(r["amount"] for r in rows)
    n_anom  = sum(1 for r in rows if r["status"] == "anomaly")
    n_warn  = sum(1 for r in rows if r["status"] == "warning")
    summary_data = [
        ["Total Pengeluaran", f"Rp {total:,.0f}"],
        ["Anomaly",           str(n_anom)],
        ["Warning",           str(n_warn)],
    ]
    summary_table = Table(summary_data, colWidths=[45*mm, 50*mm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (0, -1), colors.HexColor("#F3F4F6")),
        ("FONTNAME",    (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("GRID",        (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("PADDING",     (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 8*mm))

    # Table headers
    headers = ["Tanggal", "Kategori", "Amount (Rp)", "Note", "Status"]
    col_w   = [38*mm, 28*mm, 30*mm, 50*mm, 20*mm]
    if include_score:
        headers.append("Score")
        col_w.append(14*mm)

    table_data = [headers]
    for row in rows:
        vals = [
            row["date"], row["category"],
            f"Rp {row['amount']:,.0f}", row["note"],
            STATUS_LABEL[row["status"]],
        ]
        if include_score:
            vals.append(f"{row['score']}%" if row["score"] is not None else "-")
        table_data.append(vals)

    # Total row
    total_row_vals = ["", "TOTAL", f"Rp {total:,.0f}", "", ""]
    if include_score:
        total_row_vals.append("")
    table_data.append(total_row_vals)

    t = Table(table_data, colWidths=col_w, repeatRows=1)

    # Base style
    ts = TableStyle([
        ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, -1), 8),
        ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
        ("ALIGN",       (2, 0), (2, -1),  "RIGHT"),
        ("ALIGN",       (4, 0), (4, -1),  "CENTER"),
        ("GRID",        (0, 0), (-1, -2), 0.4, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9FAFB")]),
        ("PADDING",     (0, 0), (-1, -1), 5),
        # Total row
        ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE",   (0, -1), (-1, -1), 1, colors.HexColor("#1A1A1A")),
    ])

    # Status row colors
    status_bg = {"anomaly": "#FEE2E2", "warning": "#FEF9C3", "normal": "#DCFCE7"}
    status_col = 4
    for i, row in enumerate(rows, start=1):
        if row["status"] in status_bg:
            ts.add("BACKGROUND", (status_col, i), (status_col, i), colors.HexColor(status_bg[row["status"]]))
            txt_color = {"anomaly": "#991B1B", "warning": "#854D0E", "normal": "#166534"}[row["status"]]
            ts.add("TEXTCOLOR", (status_col, i), (status_col, i), colors.HexColor(txt_color))
            ts.add("FONTNAME",  (status_col, i), (status_col, i), "Helvetica-Bold")

    t.setStyle(ts)
    story.append(t)

    # Footer
    story.append(Spacer(1, 6*mm))
    story.append(Paragraph(
        f"Diekspor pada {datetime.now().strftime('%d %b %Y %H:%M')}",
        ParagraphStyle("footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray)
    ))

    doc.build(story)
    buf.seek(0)

    filename = f"transaksi_{MONTHS_ID[month].lower()}_{year}.pdf"
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )