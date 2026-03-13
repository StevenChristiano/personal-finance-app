import calendar
import io

from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordRequestForm
from pydantic import BaseModel, Field, EmailStr
from typing import Optional, List
from sqlalchemy.orm import Session
from datetime import datetime, timezone
import pandas as pd
import numpy as np
import pickle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import RobustScaler, LabelEncoder
from contextlib import asynccontextmanager
from dateutil.relativedelta import relativedelta

from database import get_db, User, Transaction, Category, UserModel, Income, init_db
from auth import hash_password, verify_password, create_access_token, get_current_user

# ============================================================
# STARTUP
# ============================================================
@asynccontextmanager
async def lifespan(app: FastAPI):
    init_db()  # startup
    yield

# ============================================================
# INISIALISASI APP
# ============================================================
app = FastAPI(
    title="Personal Finance Anomaly Detection API",
    description="Backend API for anomaly detection using Isolation Forest",
    version="1.0.0",
    lifespan=lifespan
)

# CORS — izinkan Next.js frontend mengakses API ini
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================
# KONFIGURASI
# ============================================================
MODEL_DIR  = "model"

THRESHOLD_ANOMALY = 0.60
THRESHOLD_WARNING = 0.50

MIN_CATEGORY = 20
MIN_GLOBAL   = 50

N_ESTIMATORS      = 100
MAX_SAMPLES       = 256
RANDOM_SEED       = 42

ANOMALY_CATEGORIES = [
    "Food", "Transport", "Lifestyle",
    "Entertainment", "Utilities", "Telecommunication", "Subscription"
]

EXCLUDED_CATEGORIES = ["Health", "Education", "Big Expense"]

# ============================================================
# LOAD GLOBAL MODEL (fallback cold start)
# ============================================================
def load_global_model():
    try:
        with open(f"{MODEL_DIR}/isolation_forest.pkl", "rb") as f:
            model = pickle.load(f)
        with open(f"{MODEL_DIR}/scaler.pkl", "rb") as f:
            scaler = pickle.load(f)
        with open(f"{MODEL_DIR}/encoder.pkl", "rb") as f:
            encoder = pickle.load(f)
        with open(f"{MODEL_DIR}/norm_params.pkl", "rb") as f:
            norm_params = pickle.load(f)
        print("✅ Global model loaded")
        return model, scaler, encoder, norm_params
    except FileNotFoundError:
        print("⚠️  Global model not found. Run train.py first.")
        return None, None, None, None

global_model, global_scaler, global_encoder, global_norm_params = load_global_model()

# ============================================================
# SCHEMAS/MODEL
# ============================================================
class RegisterRequest(BaseModel):
    email: EmailStr
    name: str
    password: str = Field(..., min_length=6)

class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user_id: int
    name: str

class TransactionCreate(BaseModel):
    amount: float = Field(..., gt=0)
    category_id: int
    note: Optional[str] = None
    timestamp: Optional[str] = None

class ColdStartStatus(BaseModel):
    is_ready: bool
    total_transactions: int
    min_global: int
    progress_global: float
    category_status: dict

class SettingsUpdate(BaseModel):
    warning_threshold: float = Field(..., ge=0.0, le=1.0)
    anomaly_threshold: float = Field(..., ge=0.0, le=1.0)

class BulkTransactionItem(BaseModel):
    amount: float = Field(..., gt=0)
    category_id: int
    note: Optional[str] = None
    timestamp: Optional[str] = None

class BulkTransactionSave(BaseModel):
    transactions: List[BulkTransactionItem]

class IncomeCreate(BaseModel):
    amount: float = Field(..., gt=0)
    source: str
    date: Optional[str] = None
    is_recurring: bool = False

# ============================================================
# HELPER: Load model user dari DB
# ============================================================
def load_user_model(user_model: UserModel):
    model       = pickle.loads(user_model.model_blob)
    scaler      = pickle.loads(user_model.scaler_blob)
    encoder     = pickle.loads(user_model.encoder_blob)
    norm_params = pickle.loads(user_model.norm_params_blob)
    return model, scaler, encoder, norm_params

# ============================================================
# HELPER: Hitung anomaly score
# ============================================================
def calculate_score(amount, category_name, timestamp, model, scaler, encoder, norm_params,
                    warning_threshold: float = THRESHOLD_WARNING,
                    anomaly_threshold: float = THRESHOLD_ANOMALY):
    hour             = timestamp.hour
    day_of_week      = timestamp.weekday()
    amount_scaled    = scaler.transform(pd.DataFrame([[amount]], columns=["amount"]))[0][0]
    category_encoded = encoder.transform([category_name])[0]
    X                = np.array([[amount_scaled, category_encoded, hour, day_of_week]])
    raw_score        = -model.score_samples(X)[0]
    min_s            = norm_params["min_score"]
    max_s            = norm_params["max_score"]
    score            = float(np.clip((raw_score - min_s) / (max_s - min_s), 0, 1))

    if score > anomaly_threshold:
        anomaly_status = "anomaly"
    elif score >= warning_threshold:
        anomaly_status = "warning"
    else:
        anomaly_status = "normal"

    return score, anomaly_status

# ============================================================
# HELPER: Retrain model user
# ============================================================
def retrain_user_model(user_id: int, db: Session, manual: bool = False):
    """
    Retrain model for a specific user.
    Selalu pakai semua transaksi non-excluded.
    Contamination dinamis hanya untuk manual retrain,
    auto-retrain pakai default 0.1.
    """
    user_model_row = db.query(UserModel).filter(UserModel.user_id == user_id).first()

    # Tentukan filter transaksi
    query = db.query(Transaction).filter(
        Transaction.user_id == user_id,
        Transaction.is_excluded == False,
    )

    transactions = query.all()

    if len(transactions) < MIN_GLOBAL:
        return False

    categories = {cat.id: cat.name for cat in db.query(Category).all()}
    data = [{
        "amount"     : t.amount,
        "category"   : categories[t.category_id],
        "hour"       : t.timestamp.hour,
        "day_of_week": t.timestamp.weekday(),
        "status"     : t.anomaly_status,
    } for t in transactions]

    df                     = pd.DataFrame(data)
    scaler                 = RobustScaler()
    encoder                = LabelEncoder()
    df["amount_scaled"]    = scaler.fit_transform(df[["amount"]])
    df["category_encoded"] = encoder.fit_transform(df["category"])
    X                      = df[["amount_scaled", "category_encoded", "hour", "day_of_week"]].values

    # Hitung contamination dinamis dari proporsi anomali user.
    # Hanya untuk manual retrain setelah personal model sudah ada
    # status anomaly di titik ini sudah dari personal model, bukan global,
    # sehingga lebih representatif sebagai estimasi kontaminasi nyata.
    if manual:
        n_anomaly     = (df["status"] == "anomaly").sum()
        raw_contam    = n_anomaly / len(df)
        contamination = float(np.clip(raw_contam, 0.01, 0.5))
        print(f"   📊 Dynamic contamination: {contamination:.3f} ({n_anomaly}/{len(df)} anomalies in training data)")
    else:
        # Auto-retrain atau first train → pakai default 10%
        # Status anomaly masih dari global model, belum reliable
        contamination = 0.1
        print(f"   📊 Default contamination: {contamination} (auto/first train)")

    model = IsolationForest(
        n_estimators=N_ESTIMATORS,
        max_samples=min(MAX_SAMPLES, len(X)),
        contamination=contamination,
        random_state=RANDOM_SEED
    )
    model.fit(X)

    raw_scores  = -model.score_samples(X)
    norm_params = {
        "min_score"    : float(raw_scores.min()),
        "max_score"    : float(raw_scores.max()),
        "contamination": contamination,
    }

    if not user_model_row:
        user_model_row = UserModel(user_id=user_id)
        db.add(user_model_row)

    user_model_row.model_blob        = pickle.dumps(model)
    user_model_row.scaler_blob       = pickle.dumps(scaler)
    user_model_row.encoder_blob      = pickle.dumps(encoder)
    user_model_row.norm_params_blob  = pickle.dumps(norm_params)
    user_model_row.transaction_count = len(transactions)
    user_model_row.is_trained        = True
    user_model_row.last_trained      = datetime.now(timezone.utc)
    db.commit()

    mode_label = "manual" if manual else "auto"
    print(f"✅ Model retrained [{mode_label}] for user {user_id} ({len(transactions)} transactions, contamination={contamination:.3f})")
    
    # Rescore semua transaksi lama dengan model baru
    rescore_all_transactions(user_id, db, model, scaler, encoder, norm_params)
    return True

# ============================================================
# HELPER: Rescore semua transaksi user dengan model terbaru
# ============================================================
def rescore_all_transactions(user_id: int, db: Session, model, scaler, encoder, norm_params):
    """
    Hitung ulang anomaly_score dan anomaly_status semua transaksi user
    menggunakan model yang baru saja ditraining.
    Dipanggil otomatis setelah setiap retrain.
    """
    user_obj  = db.query(User).filter(User.id == user_id).first()
    warning_t = user_obj.warning_threshold or THRESHOLD_WARNING
    anomaly_t = user_obj.anomaly_threshold or THRESHOLD_ANOMALY

    transactions = db.query(Transaction).filter(
        Transaction.user_id == user_id,
        Transaction.is_excluded == False,
    ).all()

    categories = {c.id: c.name for c in db.query(Category).all()}

    rescored = 0
    for t in transactions:
        cat_name = categories.get(t.category_id)
        if not cat_name or cat_name not in encoder.classes_:
            continue
        try:
            score, status = calculate_score(
                t.amount, cat_name, t.timestamp,
                model, scaler, encoder, norm_params,
                warning_threshold=warning_t,
                anomaly_threshold=anomaly_t,
            )
            t.anomaly_score  = score
            t.anomaly_status = status
            rescored += 1
        except Exception:
            continue

    db.commit()
    print(f"   🔄 Rescored {rescored}/{len(transactions)} transactions with new model")

# ============================================================
# HELPER: Cek apakah perlu auto-retrain
# ============================================================
def should_retrain(user_id: int, db: Session) -> bool:
    user_model   = db.query(UserModel).filter(UserModel.user_id == user_id).first()
    transactions = db.query(Transaction).filter(
        Transaction.user_id == user_id,
        Transaction.is_excluded == False
    ).all()
    total = len(transactions)

    if not user_model or not user_model.is_trained:
        return total >= MIN_GLOBAL

    return (total - user_model.transaction_count) >= MIN_GLOBAL


# ============================================================
# ENDPOINTS — AUTH
# ============================================================
@app.get("/")
def root():
    return {"message": "Personal Finance API", "status": "running"}

@app.post("/auth/register", status_code=201)
def register(req: RegisterRequest, db: Session = Depends(get_db)):
    if db.query(User).filter(User.email == req.email).first():
        raise HTTPException(status_code=400, detail="Email is already registered.")
    user = User(email=req.email, name=req.name, hashed_password=hash_password(req.password))
    db.add(user)
    db.commit()
    db.refresh(user)
    return {"message": "Registration Successful!", "user_id": user.id}

@app.post("/auth/login", response_model=LoginResponse)
def login(form: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == form.username).first()
    if not user or not verify_password(form.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Wrong Email or Password.")
    token = create_access_token({"sub": str(user.id)})
    return LoginResponse(access_token=token, user_id=user.id, name=user.name)

# ============================================================
# ENDPOINTS — SETTINGS
# ============================================================
@app.get("/settings")
def get_settings(current_user: User = Depends(get_current_user)):
    return {
        "warning_threshold": current_user.warning_threshold,
        "anomaly_threshold": current_user.anomaly_threshold,
    }

@app.put("/settings")
def update_settings(
    req: SettingsUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if req.warning_threshold < 0.10:
        raise HTTPException(status_code=400, detail="Warning threshold must be at least 10% (normal zone must be ≥ 10%).")
    if req.anomaly_threshold > 0.90:
        raise HTTPException(status_code=400, detail="Anomaly threshold must be at most 90% (anomaly zone must be ≥ 10%).")
    if (round(req.anomaly_threshold - req.warning_threshold, 2)) < 0.10:
        raise HTTPException(status_code=400, detail="Warning threshold and anomaly threshold must be at least 10% apart (warning zone must be ≥ 10%).")
    current_user.warning_threshold = req.warning_threshold
    current_user.anomaly_threshold = req.anomaly_threshold
    db.commit()
    return {
        "warning_threshold": current_user.warning_threshold,
        "anomaly_threshold": current_user.anomaly_threshold,
    }

# ============================================================
# ENDPOINTS — CATEGORIES
# ============================================================
@app.get("/categories")
def get_categories(db: Session = Depends(get_db)):
    return [{"id": c.id, "name": c.name, "is_excluded": c.is_excluded}
            for c in db.query(Category).all()]

# ============================================================
# ENDPOINTS — TRANSACTIONS
# ============================================================
@app.post("/transactions", status_code=201)
def create_transaction(
    req: TransactionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    category = db.query(Category).filter(Category.id == req.category_id).first()
    if not category:
        raise HTTPException(status_code=404, detail="Category Not Found.")

    ts             = datetime.fromisoformat(req.timestamp) if req.timestamp else datetime.now(timezone.utc)
    anomaly_score  = None
    anomaly_status = None
    is_excluded    = category.is_excluded

    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY

    if not is_excluded:
        user_model = db.query(UserModel).filter(
            UserModel.user_id == current_user.id,
            UserModel.is_trained == True
        ).first()

        if user_model:
            model, scaler, encoder, norm_params = load_user_model(user_model)
        elif global_model:
            model, scaler, encoder, norm_params = global_model, global_scaler, global_encoder, global_norm_params
        else:
            model = None

        if model and category.name in encoder.classes_:
            anomaly_score, anomaly_status = calculate_score(
                req.amount, category.name, ts, model, scaler, encoder, norm_params,
                warning_threshold=warning_threshold,
                anomaly_threshold=anomaly_threshold,
            )

    transaction = Transaction(
        user_id        = current_user.id,
        category_id    = req.category_id,
        amount         = req.amount,
        note           = req.note,
        timestamp      = ts,
        anomaly_score  = anomaly_score,
        anomaly_status = anomaly_status,
        is_excluded    = is_excluded
    )
    db.add(transaction)
    db.commit()
    db.refresh(transaction)

    # Auto-retrain (pakai semua data)
    if should_retrain(current_user.id, db):
        retrain_user_model(current_user.id, db, manual=False)

    return {
        "id"            : transaction.id,
        "amount"        : transaction.amount,
        "category"      : category.name,
        "note"          : transaction.note,
        "timestamp"     : transaction.timestamp,
        "anomaly_score" : transaction.anomaly_score,
        "anomaly_status": transaction.anomaly_status,
        "is_excluded"   : transaction.is_excluded,
        "message"       : "⚠️ Anomaly Detected!" if anomaly_status == "anomaly"
                          else "🔔 Unusual Transaction" if anomaly_status == "warning"
                          else "✅ Normal Transaction."
    }

@app.get("/transactions")
def get_transactions(
    month: Optional[int] = None,
    year: Optional[int] = None,
    warning_threshold: float = 0.50,
    anomaly_threshold: float = 0.60,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if month and year:
        last_day = calendar.monthrange(year, month)[1]
        query = query.filter(
            Transaction.timestamp >= datetime(year, month, 1),
            Transaction.timestamp <= datetime(year, month, last_day, 23, 59, 59)
        )
    transactions = query.order_by(Transaction.timestamp.desc()).all()
    categories   = {cat.id: cat.name for cat in db.query(Category).all()}

    def dynamic_status(score):
        if score is None: return None
        if score >= anomaly_threshold: return "anomaly"
        if score >= warning_threshold: return "warning"
        return "normal"

    return [{
        "id"            : t.id,
        "amount"        : t.amount,
        "category_id"   : t.category_id,
        "category_name" : categories.get(t.category_id, "Unknown"),
        "note"          : t.note,
        "timestamp"     : t.timestamp,
        "anomaly_score" : t.anomaly_score,
        "anomaly_status": dynamic_status(t.anomaly_score) if not t.is_excluded else None,
        "is_excluded"   : t.is_excluded
    } for t in transactions]

@app.delete("/transactions/{transaction_id}")
def delete_transaction(
    transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    transaction = db.query(Transaction).filter(
        Transaction.id == transaction_id,
        Transaction.user_id == current_user.id
    ).first()
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found.")
    db.delete(transaction)
    db.commit()
    return {"message": "Transaction Successfully Deleted."}

# ============================================================
# ENDPOINTS — STATS
# ============================================================
@app.get("/stats")
def get_stats(
    month: Optional[int] = None,
    year: Optional[int] = None,
    warning_threshold: float = 0.50,
    anomaly_threshold: float = 0.60,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(Transaction).filter(Transaction.user_id == current_user.id)
    if month and year:
        last_day = calendar.monthrange(year, month)[1]
        query = query.filter(
            Transaction.timestamp >= datetime(year, month, 1),
            Transaction.timestamp < datetime(year, month, last_day, 23, 59, 59)
        )
    transactions  = query.all()
    categories    = {cat.id: cat.name for cat in db.query(Category).all()}
    if not transactions:
        return {"total_transactions": 0, "total_amount": 0, "average_amount": 0,
                "by_category": {}, "anomaly_count": 0}

    def dynamic_status(score):
        if score is None: return None
        if score >= anomaly_threshold: return "anomaly"
        if score >= warning_threshold: return "warning"
        return "normal"

    total_amount  = sum(t.amount for t in transactions)
    anomaly_count = sum(1 for t in transactions
                        if not t.is_excluded and dynamic_status(t.anomaly_score) == "anomaly")
    by_category   = {}
    for t in transactions:
        cat_name = categories.get(t.category_id, "Unknown")
        if cat_name not in by_category:
            by_category[cat_name] = {"total": 0, "count": 0, "anomaly_count": 0}
        by_category[cat_name]["total"] += t.amount
        by_category[cat_name]["count"] += 1
        if not t.is_excluded and dynamic_status(t.anomaly_score) == "anomaly":
            by_category[cat_name]["anomaly_count"] += 1

    return {
        "total_transactions": len(transactions),
        "total_amount"      : total_amount,
        "average_amount"    : total_amount / len(transactions),
        "by_category"       : by_category,
        "anomaly_count"     : anomaly_count
    }

@app.get("/stats/monthly")
def get_monthly_stats(
    months: int = 6,
    warning_threshold: float = 0.50,
    anomaly_threshold: float = 0.60,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    result = []
    now = datetime.now(timezone.utc)

    for i in range(months - 1, -1, -1):
        target = now - relativedelta(months=i)
        m, y = target.month, target.year
        last_day = calendar.monthrange(y, m)[1]

        txs = db.query(Transaction).filter(
            Transaction.user_id == current_user.id,
            Transaction.timestamp >= datetime(y, m, 1),
            Transaction.timestamp <= datetime(y, m, last_day, 23, 59, 59)
        ).all()

        def dynamic_status(score):
            if score is None: return None
            if score >= anomaly_threshold: return "anomaly"
            if score >= warning_threshold: return "warning"
            return "normal"

        total = sum(t.amount for t in txs)
        anomaly_count = sum(1 for t in txs if not t.is_excluded and dynamic_status(t.anomaly_score) == "anomaly")

        result.append({
            "month": m,
            "year": y,
            "label": datetime(y, m, 1).strftime("%b %Y"),
            "total_amount": total,
            "transaction_count": len(txs),
            "anomaly_count": anomaly_count,
        })

    return result

# ============================================================
# ENDPOINTS — EXCEL TEMPLATE & BULK UPLOAD
# ============================================================
@app.get("/transactions/template")
def download_template(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    categories = db.query(Category).all()
    cat_names  = [c.name for c in categories]

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Transactions"

    header_fill   = PatternFill("solid", fgColor="1A1A1A")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    header_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    headers = ["date", "amount", "category", "note"]
    col_widths = [22, 18, 22, 30]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")
    examples = [
        [now_str, 50000,  "Food",          "Lunch with team"],
        [now_str, 25000,  "Transport",     "Gojek to office"],
        [now_str, 150000, "Entertainment", "Cinema ticket"],
    ]
    example_fill = PatternFill("solid", fgColor="F9FAFB")
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = example_fill
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")

    cat_list = ",".join(cat_names)
    dv = DataValidation(type="list", formula1=f'"{cat_list}"', allow_blank=True)
    dv.sqref = "C2:C10000"
    ws.add_data_validation(dv)

    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Category Reference")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 50

    ref_headers = ["Category Name", "Anomaly Scan", "Description"]
    ref_fill    = PatternFill("solid", fgColor="1A1A1A")
    for ci, hdr in enumerate(ref_headers, start=1):
        cell = ws2.cell(row=1, column=ci, value=hdr)
        cell.font      = Font(color="FFFFFF", bold=True)
        cell.fill      = ref_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws2.row_dimensions[1].height = 22

    descriptions = {
        "Food": "Daily food & beverage expenses",
        "Transport": "Commuting and travel costs",
        "Lifestyle": "Personal care and clothing",
        "Entertainment": "Leisure activities",
        "Utilities": "Household bills",
        "Telecommunication": "Phone-related expenses",
        "Subscription": "Recurring digital services",
        "Health": "Medical expenses (Not monitored)",
        "Education": "Learning costs (Not monitored)",
        "Big Expense": "Large one-time purchases (Not monitored)",
    }
    for ri, cat in enumerate(categories, start=2):
        ws2.cell(row=ri, column=1, value=cat.name).border = thin_border
        scan_cell = ws2.cell(row=ri, column=2, value="No" if cat.is_excluded else "Yes")
        scan_cell.border    = thin_border
        scan_cell.font      = Font(color="DC2626" if cat.is_excluded else "16A34A", bold=True)
        scan_cell.alignment = Alignment(horizontal="center")
        desc_cell = ws2.cell(row=ri, column=3, value=descriptions.get(cat.name, ""))
        desc_cell.border = thin_border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="transaction_template.xlsx"'},
    )


@app.post("/transactions/upload-preview")
async def upload_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded.")

    filename_lower = file.filename.lower()
    if not (filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls") or filename_lower.endswith(".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, or .csv files are accepted.")

    content = await file.read()
    buf     = io.BytesIO(content)

    try:
        if filename_lower.endswith(".csv"):
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf, engine="openpyxl", sheet_name=0)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    df.columns = [str(c).strip().lower() for c in df.columns]

    required_cols = {"date", "amount", "category"}
    missing = required_cols - set(df.columns)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}. Expected: date, amount, category, note"
        )

    df = df.dropna(subset=["date", "amount", "category"], how="all").reset_index(drop=True)
    if df.empty:
        raise HTTPException(status_code=400, detail="The file contains no valid data rows.")

    categories = {c.name.lower(): c for c in db.query(Category).all()}

    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY

    user_model_row = db.query(UserModel).filter(
        UserModel.user_id == current_user.id,
        UserModel.is_trained == True
    ).first()
    if user_model_row:
        model, scaler, encoder, norm_params = load_user_model(user_model_row)
    elif global_model:
        model, scaler, encoder, norm_params = global_model, global_scaler, global_encoder, global_norm_params
    else:
        model = None

    preview_rows = []
    for idx, row in df.iterrows():
        row_num = int(idx) + 2
        errors  = []

        try:
            amount = float(row["amount"])
            if amount <= 0:
                errors.append("Amount must be > 0")
        except (ValueError, TypeError):
            amount = 0
            errors.append("Invalid amount")

        raw_cat   = str(row["category"]).strip() if pd.notna(row.get("category")) else ""
        cat_lower = raw_cat.lower()
        if cat_lower not in categories:
            errors.append(f"Unknown category: '{raw_cat}'")
            category_obj = None
            category_id  = None
        else:
            category_obj = categories[cat_lower]
            category_id  = category_obj.id

        raw_date = row.get("date", "")
        try:
            if pd.isna(raw_date):
                ts = datetime.now()
            else:
                ts = pd.to_datetime(raw_date)
                if ts.tzinfo is not None:
                    ts = ts.replace(tzinfo=None)
                ts = ts.to_pydatetime()
        except Exception:
            ts = datetime.now()
            errors.append("Invalid date format — defaulting to now")

        note = str(row["note"]).strip() if "note" in df.columns and pd.notna(row.get("note")) else None

        anomaly_score  = None
        anomaly_status = None
        is_excluded    = category_obj.is_excluded if category_obj else False

        if not errors and not is_excluded and model and category_obj:
            if category_obj.name in encoder.classes_:
                try:
                    anomaly_score, anomaly_status = calculate_score(
                        amount, category_obj.name, ts, model, scaler, encoder, norm_params,
                        warning_threshold=warning_threshold,
                        anomaly_threshold=anomaly_threshold,
                    )
                except Exception:
                    pass

        preview_rows.append({
            "_row"          : row_num,
            "timestamp"     : ts.isoformat(),
            "amount"        : amount,
            "category_name" : category_obj.name if category_obj else raw_cat,
            "category_id"   : category_id,
            "note"          : note,
            "anomaly_score" : anomaly_score,
            "anomaly_status": anomaly_status,
            "is_excluded"   : is_excluded,
            "errors"        : errors,
        })

    return {"rows": preview_rows, "total": len(preview_rows)}


@app.post("/transactions/bulk-save", status_code=201)
def bulk_save(
    req: BulkTransactionSave,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    categories = {c.id: c for c in db.query(Category).all()}

    warning_threshold = current_user.warning_threshold or THRESHOLD_WARNING
    anomaly_threshold = current_user.anomaly_threshold or THRESHOLD_ANOMALY

    user_model_row = db.query(UserModel).filter(
        UserModel.user_id == current_user.id,
        UserModel.is_trained == True
    ).first()
    if user_model_row:
        model, scaler, encoder, norm_params = load_user_model(user_model_row)
    elif global_model:
        model, scaler, encoder, norm_params = global_model, global_scaler, global_encoder, global_norm_params
    else:
        model = None

    saved = []
    for item in req.transactions:
        category = categories.get(item.category_id)
        if not category:
            continue

        ts          = datetime.fromisoformat(item.timestamp) if item.timestamp else datetime.now(timezone.utc)
        is_excluded = category.is_excluded

        anomaly_score  = None
        anomaly_status = None
        if not is_excluded and model and category.name in encoder.classes_:
            try:
                anomaly_score, anomaly_status = calculate_score(
                    item.amount, category.name, ts, model, scaler, encoder, norm_params,
                    warning_threshold=warning_threshold,
                    anomaly_threshold=anomaly_threshold,
                )
            except Exception:
                pass

        tx = Transaction(
            user_id        = current_user.id,
            category_id    = item.category_id,
            amount         = item.amount,
            note           = item.note,
            timestamp      = ts,
            anomaly_score  = anomaly_score,
            anomaly_status = anomaly_status,
            is_excluded    = is_excluded,
        )
        db.add(tx)
        saved.append(tx)

    db.commit()
    for tx in saved:
        db.refresh(tx)

    # Auto-retrain (pakai semua data)
    if should_retrain(current_user.id, db):
        retrain_user_model(current_user.id, db, manual=False)

    return {"saved": len(saved), "message": f"{len(saved)} transactions saved successfully."}


# ============================================================
# ENDPOINTS — RETRAIN (MANUAL)
# ============================================================
@app.post("/retrain")
def manual_retrain(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    User-triggered retrain. Requires at least MIN_GLOBAL non-excluded transactions.
    Filters out anomaly transactions if user already has a personal model,
    so the model only learns from normal + warning patterns.
    """
    total = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.is_excluded == False
    ).count()

    if total < MIN_GLOBAL:
        raise HTTPException(
            status_code=400,
            detail=f"Not enough data to train. Need at least {MIN_GLOBAL} transactions, currently have {total}."
        )

    success = retrain_user_model(current_user.id, db, manual=True)
    if not success:
        raise HTTPException(status_code=500, detail="Retrain failed. Please try again.")

    user_model  = db.query(UserModel).filter(UserModel.user_id == current_user.id).first()
    norm_params = pickle.loads(user_model.norm_params_blob)

    # Hitung kategori yang datanya masih kurang
    all_transactions = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.is_excluded == False,
    ).all()
    categories_map  = {c.id: c.name for c in db.query(Category).filter(Category.is_excluded == False).all()}
    cat_counts: dict = {}
    for t in all_transactions:
        cat_name = categories_map.get(t.category_id)
        if cat_name:
            cat_counts[cat_name] = cat_counts.get(cat_name, 0) + 1

    low_data_categories = [
        {"category": cat, "count": count, "min_required": MIN_CATEGORY}
        for cat, count in cat_counts.items()
        if count < MIN_CATEGORY
    ]
    low_data_categories.sort(key=lambda x: x["count"])

    return {
        "message"          : "Personal model successfully trained!",
        "transaction_count": user_model.transaction_count,
        "last_trained"     : user_model.last_trained,
        "contamination"    : norm_params.get("contamination", 0.1),
        "low_data_categories": low_data_categories,
    }


# ============================================================
# ENDPOINTS — COLD START & MODEL STATUS
# ============================================================
@app.get("/cold-start-status", response_model=ColdStartStatus)
def cold_start_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    transactions = db.query(Transaction).filter(
        Transaction.user_id == current_user.id,
        Transaction.is_excluded == False
    ).all()
    total      = len(transactions)
    categories = {cat.id: cat.name for cat in db.query(Category).filter(Category.is_excluded == False).all()}
    cat_counts = {}
    for t in transactions:
        cat_name = categories.get(t.category_id, "Unknown")
        cat_counts[cat_name] = cat_counts.get(cat_name, 0) + 1

    return ColdStartStatus(
        is_ready           = total >= MIN_GLOBAL,
        total_transactions = total,
        min_global         = MIN_GLOBAL,
        progress_global    = min(total / MIN_GLOBAL * 100, 100),
        category_status    = {
            cat: {"count": cat_counts.get(cat, 0), "min_required": MIN_CATEGORY,
                  "is_ready": cat_counts.get(cat, 0) >= MIN_CATEGORY}
            for cat in categories.values()
        }
    )

@app.get("/model-status")
def model_status(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    user_model = db.query(UserModel).filter(UserModel.user_id == current_user.id).first()
    if user_model and user_model.is_trained:
        return {
            "status"            : "personal",
            "message"           : "Using your Personal Model",
            "transaction_count" : user_model.transaction_count,
            "last_trained"      : user_model.last_trained,
        }
    elif global_model:
        return {"status": "global", "message": "Using global model (cold start)"}
    return {"status": "not_loaded", "message": "Model not yet available."}


# ============================================================
# ENDPOINTS — INCOME
# ============================================================
@app.post("/income", status_code=201)
def create_income(
    req: IncomeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    date = datetime.fromisoformat(req.date) if req.date else datetime.now(timezone.utc)
    income = Income(
        user_id      = current_user.id,
        amount       = req.amount,
        source       = req.source,
        date         = date,
        is_recurring = req.is_recurring,
    )
    db.add(income)
    db.commit()
    db.refresh(income)
    return {
        "id"          : income.id,
        "amount"      : income.amount,
        "source"      : income.source,
        "date"        : income.date,
        "is_recurring": income.is_recurring,
    }
 
@app.get("/income")
def get_income(
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Return pemasukan user. Kalau month+year diberikan, filter by bulan.
    Recurring income dari bulan sebelumnya otomatis di-generate kalau belum ada.
    """
    # Auto-generate recurring income untuk bulan ini kalau belum ada
    if month and year:
        _ensure_recurring_income(current_user.id, month, year, db)
 
    query = db.query(Income).filter(Income.user_id == current_user.id)
    if month and year:
        last_day = calendar.monthrange(year, month)[1]
        query = query.filter(
            Income.date >= datetime(year, month, 1),
            Income.date <= datetime(year, month, last_day, 23, 59, 59)
        )
    incomes = query.order_by(Income.date.desc()).all()
    return [{
        "id"          : i.id,
        "amount"      : i.amount,
        "source"      : i.source,
        "date"        : i.date,
        "is_recurring": i.is_recurring,
    } for i in incomes]
 
@app.delete("/income/{income_id}")
def delete_income(
    income_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    income = db.query(Income).filter(
        Income.id == income_id,
        Income.user_id == current_user.id
    ).first()
    if not income:
        raise HTTPException(status_code=404, detail="Income not found.")
    db.delete(income)
    db.commit()
    return {"message": "Income deleted."}
 
@app.get("/balance")
def get_balance(
    month: Optional[int] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    now = datetime.now(timezone.utc)
    m   = month or now.month
    y   = year  or now.year
 
    # Auto-generate recurring income bulan ini
    _ensure_recurring_income(current_user.id, m, y, db)
 
    last_day = calendar.monthrange(y, m)[1]
 
    # Total all-time
    all_incomes      = db.query(Income).filter(Income.user_id == current_user.id).all()
    all_transactions = db.query(Transaction).filter(
        Transaction.user_id    == current_user.id,
        Transaction.is_excluded == False,
    ).all()
    total_income  = sum(i.amount for i in all_incomes)
    total_expense = sum(t.amount for t in all_transactions)
 
    # Bulanan
    monthly_incomes = db.query(Income).filter(
        Income.user_id == current_user.id,
        Income.date    >= datetime(y, m, 1),
        Income.date    <= datetime(y, m, last_day, 23, 59, 59),
    ).all()
    monthly_transactions = db.query(Transaction).filter(
        Transaction.user_id    == current_user.id,
        Transaction.is_excluded == False,
        Transaction.timestamp  >= datetime(y, m, 1),
        Transaction.timestamp  <= datetime(y, m, last_day, 23, 59, 59),
    ).all()
    monthly_income  = sum(i.amount for i in monthly_incomes)
    monthly_expense = sum(t.amount for t in monthly_transactions)
 
    return {
        "total_balance"  : total_income - total_expense,
        "monthly_balance" : monthly_income - monthly_expense,
        "total_income"   : total_income,
        "total_expense"  : total_expense,
        "monthly_income" : monthly_income,
        "monthly_expense": monthly_expense,
        "month"          : m,
        "year"           : y,
    }

def _ensure_recurring_income(user_id: int, month: int, year: int, db: Session):
    """
    Auto-generate recurring income untuk bulan tertentu
    berdasarkan recurring income dari bulan sebelumnya,
    kalau belum ada di bulan tersebut.
    """
    last_day = calendar.monthrange(year, month)[1]
    start    = datetime(year, month, 1)
    end      = datetime(year, month, last_day, 23, 59, 59)
 
    # Cari semua recurring income yang sudah ada bulan ini
    existing_sources = {
        i.source for i in db.query(Income).filter(
            Income.user_id    == user_id,
            Income.is_recurring == True,
            Income.date       >= start,
            Income.date       <= end,
        ).all()
    }
 
    # Cari recurring income dari bulan-bulan sebelumnya
    prev_recurring = db.query(Income).filter(
        Income.user_id      == user_id,
        Income.is_recurring == True,
        Income.date         < start,
    ).all()
 
    # Group by source — ambil yang terbaru per source
    latest: dict = {}
    for i in prev_recurring:
        if i.source not in latest or i.date > latest[i.source].date:
            latest[i.source] = i
 
    # Generate yang belum ada
    for source, inc in latest.items():
        if source not in existing_sources:
            db.add(Income(
                user_id      = user_id,
                amount       = inc.amount,
                source       = source,
                date         = datetime(year, month, inc.date.day if inc.date.day <= last_day else last_day),
                is_recurring = True,
            ))
    db.commit()
 

# ============================================================
# RUN SERVER
# ============================================================
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)